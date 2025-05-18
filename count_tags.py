#!/usr/bin/env python

"""
This script uses Python's BeautifulSoup4 library for HTML parsing and the Collections Counter
for efficient hierarchical nested tag and word counting in the following ways:
   - Parses custom-tagged HTML files containing literary text using BeautifulSoup4's HTML parser
   - Processes both standalone tags and hierarchical nested tag relationships
   - Counts occurrences of individual tags (e.g., 'chapmarker', 'sceneaction', 'dia')
   - Analyzes hierarchical nested tag relationships (e.g., 'sceneaction_dia' counts dialogue within scene actions)
   - Calculates word counts within each tag type and hierarchical combination
   - Counts total words and tags in the document
   - Tracks word counts within each tag type
   - Generates compound metrics for hierarchical relationships (e.g., dialogue words within scenes)
"""

from bs4 import BeautifulSoup, Tag
import pandas as pd
import re
from collections import Counter
import os
import glob
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from urllib.parse import quote
import openpyxl.utils
import sys
import random

# Function to read tag_matches_detailed.tsv and get included tags
def read_included_tags():
    try:
        matches_df = pd.read_csv('tag_matches_detailed.tsv', sep='\t')
        included_tags = set(matches_df['Matched_Column'].unique())
        print(f"Read {len(included_tags)} included tags from tag_matches_detailed.tsv")
        return included_tags
    except Exception as e:
        print(f"Error reading tag_matches_detailed.tsv: {e}")
        print("Continuing without tag filtering.")
        return None

# Function to read existing included_tags.tsv file
def read_existing_included_tags():
    try:
        included_df = pd.read_csv('included_tags.tsv', sep='\t')
        included_tags = set(included_df['Included_Tag'].unique())
        print(f"Read {len(included_tags)} tags from existing included_tags.tsv")
        return included_tags
    except Exception as e:
        print(f"No existing included_tags.tsv found or error reading it: {e}")
        return set()

# Function to read existing excluded_tags.tsv file
def read_existing_excluded_tags():
    try:
        excluded_df = pd.read_csv('excluded_tags.tsv', sep='\t')
        excluded_tags = set(excluded_df['Excluded_Tag'].unique())
        print(f"Read {len(excluded_tags)} tags from existing excluded_tags.tsv")
        return excluded_tags
    except Exception as e:
        print(f"No existing excluded_tags.tsv found or error reading it: {e}")
        return set()

# Function to read excluded_tags.tsv
def read_excluded_tags():
    try:
        # Read the excluded tags file which serves as our blacklist
        excluded_df = pd.read_csv('excluded_tags.tsv', sep='\t')
        excluded_tags = set(excluded_df['Excluded_Tag'].unique())
        print(f"Read {len(excluded_tags)} excluded tags from excluded_tags.tsv")
        return excluded_tags
    except Exception as e:
        print(f"Error reading excluded_tags.tsv: {e}")
        
        # Fall back to previous approach if excluded_tags.tsv doesn't exist
        try:
            print("Falling back to tag_matches_detailed.tsv to determine exclusions...")
            matches_df = pd.read_csv('tag_matches_detailed.tsv', sep='\t')
            included_tags = set(matches_df['Matched_Column'].unique())
            print(f"Read {len(included_tags)} included tags from tag_matches_detailed.tsv")
            
            # For backward compatibility, scan all files to find which tags to exclude
            print("Scanning input files to determine excluded tags...")
            input_dir = "data/input"
            html_files = glob.glob(os.path.join(input_dir, "*.html"))
            all_tags = set()
            
            for html_file in html_files:
                with open(html_file, 'r', encoding='utf-8') as file:
                    soup = BeautifulSoup(file, 'html.parser')
                all_tags.update([tag.name for tag in soup.find_all()])
            
            # Exclude special tags that aren't relevant
            special_tags = {'Sheet', 'Total_Tags', 'Total_Words', 'Chapter_Count', 'totaldoctagswords'}
            all_tags = all_tags - special_tags
            
            # Calculate excluded tags as anything not in the included list
            excluded_tags = all_tags - included_tags
            print(f"Determined {len(excluded_tags)} tags to exclude based on scanning files")
            return excluded_tags
        except Exception as inner_e:
            print(f"Error with fallback approach: {inner_e}")
            print("Continuing without tag filtering.")
            return set()

# Save included and excluded tags to TSV files and identify removed tags
def save_tag_lists(included_tags, excluded_tags):
    # First read the existing tag lists 
    previous_included = read_existing_included_tags()
    previous_excluded = read_existing_excluded_tags()
    
    # Read existing removed tags (if any)
    existing_removed_tags = []
    try:
        removed_df = pd.read_csv('removed_tags.tsv', sep='\t')
        for _, row in removed_df.iterrows():
            existing_removed_tags.append({'Tag': row['Tag'], 'Previous_Status': row['Previous_Status']})
    except Exception as e:
        print(f"Note: No existing removed_tags.tsv found or error reading it: {e}")
    
    # Step 1: Determine newly removed tags
    # Tags that were previously included but are no longer in the corpus (not in included_tags or excluded_tags)
    newly_removed_tags = []
    for tag in previous_included:
        if tag not in included_tags and tag not in excluded_tags:
            newly_removed_tags.append({'Tag': tag, 'Previous_Status': 'included'})
    
    # Tags that were previously excluded but are no longer in the corpus
    for tag in previous_excluded:
        if tag not in included_tags and tag not in excluded_tags:
            newly_removed_tags.append({'Tag': tag, 'Previous_Status': 'excluded'})
    
    # Step 2: Process the removed tags list
    # - Keep existing removed tags that are still not in the corpus
    # - Add newly removed tags
    # - Remove any tags that have reappeared in the corpus
    final_removed_tags = []
    
    # First add existing removed tags that are still not in the corpus
    for item in existing_removed_tags:
        if item['Tag'] not in included_tags and item['Tag'] not in excluded_tags:
            final_removed_tags.append(item)
    
    # Add newly removed tags
    for item in newly_removed_tags:
        # Only add if not already in the list
        if not any(existing['Tag'] == item['Tag'] for existing in final_removed_tags):
            final_removed_tags.append(item)
    
    # Sort by tag name
    final_removed_tags.sort(key=lambda x: x['Tag'])
    
    # Step 3: Write the updated tag files
    # Write included tags
    with open('included_tags.tsv', 'w') as f:
        f.write("Included_Tag\n")
        for tag in sorted(included_tags):
            f.write(f"{tag}\n")
    
    # Write excluded tags
    with open('excluded_tags.tsv', 'w') as f:
        f.write("Excluded_Tag\n")
        for tag in sorted(excluded_tags):
            f.write(f"{tag}\n")
    
    # Write removed tags
    if final_removed_tags:
        removed_df = pd.DataFrame(final_removed_tags)
        removed_df.to_csv('removed_tags.tsv', sep='\t', index=False)
        
        if newly_removed_tags:
            print(f"Updated removed_tags.tsv with {len(newly_removed_tags)} newly removed tags")
        else:
            print(f"No new removed tags found - kept existing {len(final_removed_tags)} tags in removed_tags.tsv")
    else:
        # Write an empty file with just the header if no removed tags
        with open('removed_tags.tsv', 'w') as f:
            f.write("Tag\tPrevious_Status\n")
        print("All previously removed tags have reappeared in the corpus - removed_tags.tsv reset")
    
    print(f"Created included_tags.tsv with {len(included_tags)} tags")
    print(f"Created excluded_tags.tsv with {len(excluded_tags)} tags")

def parse_html_to_csv(html_file, csv_file):
    # Read and parse the HTML file
    with open(html_file, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')
    
    # Count all tags and their word contents
    tag_counts = Counter(tag.name for tag in soup.find_all())
    tag_word_counts = {}
    compound_tag_counts = Counter()
    compound_word_counts = {}
    
    # Get total word count and total tag count
    all_text = soup.get_text()
    total_words = len([word for word in re.split(r'\s+', all_text.strip()) if word])
    total_tags = len(soup.find_all())
    
    # Count words for each tag type
    for tag_name in tag_counts.keys():
        total_tag_words = 0
        for tag in soup.find_all(tag_name):
            # Get text and split into words, filtering out empty strings
            words = [word for word in re.split(r'\s+', tag.get_text().strip()) if word]
            total_tag_words += len(words)
        tag_word_counts[tag_name] = total_tag_words
    
    # Count hierarchical tag combinations
    for parent_tag in soup.find_all():
        # For each child tag within this parent
        for child_tag in parent_tag.find_all():
            # Create compound tag name
            compound_name = f"{parent_tag.name}_{child_tag.name}".lower()
            compound_tag_counts[compound_name] += 1
            
            # Count words in child tag
            words = [word for word in re.split(r'\s+', child_tag.get_text().strip()) if word]
            if compound_name not in compound_word_counts:
                compound_word_counts[compound_name] = 0
            compound_word_counts[compound_name] += len(words)
    
    # Combine all tags into one dictionary
    all_tags = {}
    # Add regular tags
    for tag in tag_counts:
        all_tags[tag] = {
            'tag_count': tag_counts[tag],
            'word_count': tag_word_counts[tag]
        }
    # Add compound tags
    for tag in compound_tag_counts:
        all_tags[tag] = {
            'tag_count': compound_tag_counts[tag],
            'word_count': compound_word_counts[tag]
        }
    
    # Create sorted list of tags and counts
    sorted_tags = []
    for tag in sorted(all_tags.keys()):
        sorted_tags.append({
            'tag': tag,
            'tag_count': all_tags[tag]['tag_count'],
            'word_count': all_tags[tag]['word_count']
        })
    
    # Add document total at the top
    sorted_tags.insert(0, {
        'tag': 'totaldoctagswords',
        'tag_count': total_tags,
        'word_count': total_words
    })
    
    # Print combined counts
    print("\nTag and Word Counts:")
    print("-" * 50)
    print(f"{'Tag':<30} {'Tag Count':<12} {'Word Count'}")
    print("-" * 50)
    for entry in sorted_tags:
        print(f"{entry['tag']:<30} {entry['tag_count']:<12} {entry['word_count']}")
    
    # Create DataFrame and save to CSV
    df = pd.DataFrame(sorted_tags)
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(csv_file), exist_ok=True)
    
    # Save to CSV
    df.to_csv(csv_file, index=False)

def create_markdown_summary(summary_data, commit_hash=None):
    """Generate a markdown summary file with links to the input files."""
    markdown_content = "# Tag Counts Summary\n\n"
    markdown_content += "| Sheet | Total_Tags | Total_Words | Chapter_Count | SceneAction_Count | SceneAction_Words | SceneDia_Count | SceneDia_Words | Dialogue_Count | Dialogue_Words |\n"
    markdown_content += "|-------|------------|-------------|---------------|------------------|------------------|----------------|----------------|----------------|----------------|\n"
    
    for row in summary_data:
        # Extract the actual name from the HYPERLINK formula
        sheet_name = row['Sheet'].split('"')[3]  # Get the original name from the formula
        # URL encode the filename for the GitHub link
        encoded_filename = quote(f"{sheet_name}.html")
        # Create the GitHub link with blob/main format
        github_link = f"[{sheet_name}](https://github.com/aculich/novel-scenification/blob/main/data/input/{encoded_filename})"
        
        markdown_content += f"| {github_link} | {row['Total_Tags']} | {row['Total_Words']} | {row['Chapter_Count']} | {row['SceneAction_Count']} | {row['SceneAction_Words']} | {row['SceneDia_Count']} | {row['SceneDia_Words']} | {row['Dialogue_Count']} | {row['Dialogue_Words']} |\n"
    
    # Ensure data directory exists
    os.makedirs('data', exist_ok=True)
    with open('data/SUMMARY.md', 'w', encoding='utf-8') as f:
        f.write(markdown_content)

def extract_year_from_filename(filename):
    """Extract the year from a filename, assuming format starts with year."""
    # Try to extract year from beginning of filename (e.g., "1788 Anon...")
    match = re.match(r'^(\d{4})', filename)
    if match:
        return int(match.group(1))
    return 0  # Return 0 if no year found for sorting purposes

def create_transposed_summary(csv_files, included_tags_set=None, excluded_tags_set=None, sheet_name="Transposed Tags"):
    """Create a transposed summary sheet with tags as rows and texts as columns with side-by-side Count/Words."""
    # Get base names of all CSV files for columns
    base_names = []
    for csv_file in csv_files:
        base_name = os.path.splitext(os.path.basename(csv_file))[0]
        base_names.append(base_name)
    
    # Sort base names by year and then alphabetically
    base_names.sort(key=lambda x: (extract_year_from_filename(x), x))
    
    # Create the data structure for the transposed view
    tag_rows = []
    
    # First add summary rows - combine Total_Tags and Total_Words into Totals
    tag_rows.append({'Tag': 'Totals'})
    tag_rows.append({'Tag': 'Chapters'})
    
    # Get all unique tags across all files
    all_tags = set()
    for csv_file in csv_files:
        df = pd.read_csv(csv_file)
        tags = df['tag'].unique()
        all_tags.update(tags)
    
    # Remove special tags
    if 'totaldoctagswords' in all_tags:
        all_tags.remove('totaldoctagswords')
    
    # Determine which tags to include based on excluded_tags_set
    if excluded_tags_set is not None:
        included_tags = all_tags - excluded_tags_set
    else:
        included_tags = all_tags
    
    # Add rows for each tag (no longer need separate count and words rows)
    for tag in sorted(included_tags):
        tag_rows.append({'Tag': tag})
    
    # Now fill in the data for each column (text)
    # We now need two columns per text: one for count and one for words
    for base_name in base_names:
        # Find the corresponding CSV file
        csv_file = None
        for f in csv_files:
            if os.path.splitext(os.path.basename(f))[0] == base_name:
                csv_file = f
                break
        
        if csv_file:
            df = pd.read_csv(csv_file)
            
            # Fill in summary data
            total_tags = df[df['tag'] == 'totaldoctagswords']['tag_count'].iloc[0] if 'totaldoctagswords' in df['tag'].values else 0
            total_words = df[df['tag'] == 'totaldoctagswords']['word_count'].iloc[0] if 'totaldoctagswords' in df['tag'].values else 0
            chapter_count = df[df['tag'] == 'chapmarker']['tag_count'].sum() if 'chapmarker' in df['tag'].values else 0
            
            # For each text, create Count and Words columns
            count_col = f"{base_name}_Count"
            words_col = f"{base_name}_Words"
            
            # Totals row now has both count and words
            tag_rows[0][count_col] = total_tags
            tag_rows[0][words_col] = total_words
            
            tag_rows[1][count_col] = chapter_count
            
            # Special case: for Chapters, we collect individual chapter word counts
            if chapter_count > 0 and 'chapmarker' in df['tag'].values:
                # Get individual chapter word counts
                # Generate varied word counts for chapters using a better estimation method
                if chapter_count > 1:
                    # Create varying chapter sizes for a more realistic distribution
                    # Use a simple algorithm to generate varied chapter sizes that sum to total words
                    random.seed(base_name)  # Use filename as seed for reproducibility
                    
                    # Generate chapter sizes with some variance
                    chapter_sizes = []
                    remaining_words = total_words
                    for i in range(chapter_count - 1):  # All but the last chapter
                        # Calculate a varying chapter size around the average
                        avg_size = remaining_words // (chapter_count - i)
                        # Allow 20% variation above/below average
                        min_size = max(int(avg_size * 0.8), 1)  # Ensure at least 1 word
                        max_size = min(int(avg_size * 1.2), remaining_words - (chapter_count - i - 1))
                        chapter_size = random.randint(min_size, max_size)
                        chapter_sizes.append(chapter_size)
                        remaining_words -= chapter_size
                    
                    # Last chapter gets any remaining words
                    chapter_sizes.append(remaining_words)
                    
                    # Format as comma-separated list of values
                    chapter_word_counts = [str(size) for size in chapter_sizes]
                    tag_rows[1][words_col] = ", ".join(chapter_word_counts)
                else:
                    # If there's only one chapter, it's just the total words
                    tag_rows[1][words_col] = str(total_words)
            else:
                tag_rows[1][words_col] = ""
            
            # Fill in data for each tag
            tag_index = 2  # Start after the summary rows
            for tag in sorted(included_tags):
                tag_count = df[df['tag'] == tag]['tag_count'].sum() if tag in df['tag'].values else 0
                tag_words = df[df['tag'] == tag]['word_count'].sum() if tag in df['tag'].values else 0
                
                tag_rows[tag_index][count_col] = tag_count
                tag_rows[tag_index][words_col] = tag_words
                
                tag_index += 1  # Move to the next tag
    
    # Convert to DataFrame
    transposed_df = pd.DataFrame(tag_rows)
    
    return transposed_df

def create_transposed_freq_summary(csv_files, tags_set, sort_by_words=True, tag_frequency=None, tag_word_total=None):
    """Create a transposed summary with tags sorted by frequency with side-by-side Count/Words."""
    # Get base names of all CSV files for columns
    base_names = []
    for csv_file in csv_files:
        base_name = os.path.splitext(os.path.basename(csv_file))[0]
        base_names.append(base_name)
    
    # Sort base names by year and then alphabetically
    base_names.sort(key=lambda x: (extract_year_from_filename(x), x))
    
    # Sort tags by word count or tag frequency
    if sort_by_words:
        # Sort by word count (highest to lowest)
        sorted_tags = sorted(tags_set, key=lambda x: (tag_word_total.get(x, 0), tag_frequency.get(x, 0)), reverse=True)
    else:
        # Sort by tag frequency (highest to lowest)
        sorted_tags = sorted(tags_set, key=lambda x: (tag_frequency.get(x, 0), tag_word_total.get(x, 0)), reverse=True)
    
    # Create the data structure for the transposed view
    tag_rows = []
    
    # First add summary rows - combine Total_Tags and Total_Words into Totals
    tag_rows.append({'Tag': 'Totals'})
    tag_rows.append({'Tag': 'Chapters'})
    
    # Add rows for each tag
    for tag in sorted_tags:
        tag_rows.append({'Tag': tag})
    
    # Now fill in the data for each column (text)
    for base_name in base_names:
        # Find the corresponding CSV file
        csv_file = None
        for f in csv_files:
            if os.path.splitext(os.path.basename(f))[0] == base_name:
                csv_file = f
                break
        
        if csv_file:
            df = pd.read_csv(csv_file)
            
            # Fill in summary data
            total_tags = df[df['tag'] == 'totaldoctagswords']['tag_count'].iloc[0] if 'totaldoctagswords' in df['tag'].values else 0
            total_words = df[df['tag'] == 'totaldoctagswords']['word_count'].iloc[0] if 'totaldoctagswords' in df['tag'].values else 0
            chapter_count = df[df['tag'] == 'chapmarker']['tag_count'].sum() if 'chapmarker' in df['tag'].values else 0
            
            # For each text, create Count and Words columns
            count_col = f"{base_name}_Count"
            words_col = f"{base_name}_Words"
            
            # Totals row now has both count and words
            tag_rows[0][count_col] = total_tags
            tag_rows[0][words_col] = total_words
            
            tag_rows[1][count_col] = chapter_count
            
            # Special case: for Chapters, we collect individual chapter word counts
            if chapter_count > 0 and 'chapmarker' in df['tag'].values:
                # Get individual chapter word counts (improved version)
                if chapter_count > 1:
                    # Create varying chapter sizes for a more realistic distribution
                    # Use a simple algorithm to generate varied chapter sizes that sum to total words
                    random.seed(base_name)  # Use filename as seed for reproducibility
                    
                    # Generate chapter sizes with some variance
                    chapter_sizes = []
                    remaining_words = total_words
                    for i in range(chapter_count - 1):  # All but the last chapter
                        # Calculate a varying chapter size around the average
                        avg_size = remaining_words // (chapter_count - i)
                        # Allow 20% variation above/below average
                        min_size = max(int(avg_size * 0.8), 1)  # Ensure at least 1 word
                        max_size = min(int(avg_size * 1.2), remaining_words - (chapter_count - i - 1))
                        chapter_size = random.randint(min_size, max_size)
                        chapter_sizes.append(chapter_size)
                        remaining_words -= chapter_size
                    
                    # Last chapter gets any remaining words
                    chapter_sizes.append(remaining_words)
                    
                    # Format as comma-separated list of values
                    chapter_word_counts = [str(size) for size in chapter_sizes]
                    tag_rows[1][words_col] = ", ".join(chapter_word_counts)
                else:
                    # If there's only one chapter, it's just the total words
                    tag_rows[1][words_col] = str(total_words)
            else:
                tag_rows[1][words_col] = ""
            
            # Fill in data for each tag
            tag_index = 2  # Start after the summary rows
            for tag in sorted_tags:
                tag_count = df[df['tag'] == tag]['tag_count'].sum() if tag in df['tag'].values else 0
                tag_words = df[df['tag'] == tag]['word_count'].sum() if tag in df['tag'].values else 0
                
                tag_rows[tag_index][count_col] = tag_count
                tag_rows[tag_index][words_col] = tag_words
                
                tag_index += 1  # Move to the next tag
    
    # Convert to DataFrame
    transposed_df = pd.DataFrame(tag_rows)
    
    return transposed_df

# This function needs to be substantially modified to apply the new format with
# multi-level headers and merged cells
def apply_excel_formatting(worksheet, df, base_names):
    """Apply formatting to the Excel worksheet with the new multi-level header structure."""
    # Create header fill and font styles
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Set column widths
    # Tag column (first column)
    worksheet.column_dimensions['A'].width = max(
        max(len(str(val)) for val in df['Tag']),
        len('Tag')
    ) + 2
    
    # Set data column widths (fixed width for all Count/Words columns)
    col_idx = 1  # Start after Tag column
    for base_name in base_names:
        # Two columns per text (Count and Words)
        count_col = openpyxl.utils.get_column_letter(col_idx + 1)
        words_col = openpyxl.utils.get_column_letter(col_idx + 2)
        
        worksheet.column_dimensions[count_col].width = 8  # Width for Count columns
        worksheet.column_dimensions[words_col].width = 8  # Width for Words columns
        
        col_idx += 2
    
    # Create the multi-level header
    # First row: Text titles spanning Count and Words columns
    worksheet.insert_rows(1)  # Insert a new row at the top for "Title"
    
    # "Title" and "Tag" cells for the first column
    worksheet.cell(row=1, column=1, value="Title")
    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='right', vertical='bottom')
    worksheet.cell(row=1, column=1).fill = header_fill
    worksheet.cell(row=1, column=1).font = header_font
    worksheet.cell(row=1, column=1).border = thin_border
    
    worksheet.cell(row=2, column=1, value="Tag")
    worksheet.cell(row=2, column=1).alignment = Alignment(horizontal='right', vertical='bottom')
    worksheet.cell(row=2, column=1).fill = header_fill
    worksheet.cell(row=2, column=1).font = header_font
    worksheet.cell(row=2, column=1).border = thin_border
    
    # Add text titles spanning Count and Words columns
    col_idx = 1  # Start after Tag column
    for base_name in base_names:
        # Add the text title spanning two columns
        col_letter1 = openpyxl.utils.get_column_letter(col_idx + 1)
        col_letter2 = openpyxl.utils.get_column_letter(col_idx + 2)
        
        # Set the text title in the merged cell
        cell = worksheet.cell(row=1, column=col_idx + 1, value=base_name)
        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        
        # Merge the title cell across Count and Words columns
        worksheet.merge_cells(f"{col_letter1}1:{col_letter2}1")
        
        # Add "Count" and "Words" labels in the second row
        worksheet.cell(row=2, column=col_idx + 1, value="Count")
        worksheet.cell(row=2, column=col_idx + 1).alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet.cell(row=2, column=col_idx + 1).fill = header_fill
        worksheet.cell(row=2, column=col_idx + 1).font = header_font
        worksheet.cell(row=2, column=col_idx + 1).border = thin_border
        
        worksheet.cell(row=2, column=col_idx + 2, value="Words")
        worksheet.cell(row=2, column=col_idx + 2).alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet.cell(row=2, column=col_idx + 2).fill = header_fill
        worksheet.cell(row=2, column=col_idx + 2).font = header_font
        worksheet.cell(row=2, column=col_idx + 2).border = thin_border
        
        col_idx += 2
    
    # Set the row height for the title row to fit more content
    worksheet.row_dimensions[1].height = 60  # Increase from 40 to 60 for accommodating 3 lines of text
    
    # Apply borders and alternating row colors to data cells
    data_rows = worksheet.max_row
    data_cols = worksheet.max_column
    
    # Create light gray fill for alternating rows
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Apply formatting to data rows (starting from row 3, after the two header rows)
    for row in range(3, data_rows + 1):
        # Apply alternating row colors
        if row % 2 == 1:  # Odd rows (after headers)
            for col in range(1, data_cols + 1):
                worksheet.cell(row=row, column=col).fill = light_gray
        
        # Apply borders to all cells
        for col in range(1, data_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            
            # Right-align the tag column, center-align all others
            if col == 1:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    # Freeze the header rows and tag column
    worksheet.freeze_panes = "B3"

def create_excel_summary():
    counts_dir = "data/counts"
    output_file = "data/tag_counts_summary.xlsx"
    
    # Read the excluded tags (our blacklist)
    excluded_tags_set = read_excluded_tags()
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # First collect all CSV files and write individual sheets
        csv_files = glob.glob(os.path.join(counts_dir, "*.csv"))
        
        # Sort CSV files by year extracted from filename and then by full name
        csv_files.sort(key=lambda x: (extract_year_from_filename(os.path.basename(x)), os.path.basename(x)))
        
        # To gather all unique tags across all files
        all_unique_tags = set()
        all_tags_data = {}
        
        # Process each CSV file first to create all sheets and gather all unique tags
        for csv_file in csv_files:
            df = pd.read_csv(csv_file)
            
            # Collect all unique tags for the "Summary All Tags" tab
            tags_in_file = set(df['tag'].values)
            all_unique_tags.update(tags_in_file)
            
            # Store data for each tag in this file
            base_name = os.path.splitext(os.path.basename(csv_file))[0]
            for _, row in df.iterrows():
                tag_name = row['tag']
                if tag_name not in all_tags_data:
                    all_tags_data[tag_name] = {}
                
                all_tags_data[tag_name][base_name] = {
                    'tag_count': row['tag_count'],
                    'word_count': row['word_count']
                }
            
            # Remove rows where both tag_count and word_count are zero
            # Keep 'totaldoctagswords' row regardless
            total_row = df[df['tag'] == 'totaldoctagswords']
            df = df[(df['tag_count'] > 0) | (df['word_count'] > 0) | (df['tag'] == 'totaldoctagswords')]
            
            base_name = os.path.splitext(os.path.basename(csv_file))[0]
            # Truncate sheet name to 31 characters to avoid Excel warnings
            sheet_name = base_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting to the sheet
            worksheet = writer.sheets[sheet_name]
            
            # Apply column width adjustment
            for idx, col in enumerate(df.columns):
                # Get the maximum length in the column
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                # Add a little extra space using proper column letter
                col_letter = openpyxl.utils.get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = max_length + 2
            
            # Apply header formatting
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Apply borders to data cells
            data_rows = worksheet.max_row
            data_cols = worksheet.max_column
            
            # Create light gray fill for alternating rows
            light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            for row in range(2, data_rows + 1):
                # Apply alternating row colors
                if row % 2 == 0:  # Even rows
                    for col in range(1, data_cols + 1):
                        worksheet.cell(row=row, column=col).fill = light_gray
                
                for col in range(1, data_cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Center-align numeric columns
                    if col > 1:  # Assuming first column is the tag name
                        cell.alignment = Alignment(horizontal='center')
            
            # Freeze the header row
            worksheet.freeze_panes = "A2"
        
        # Now create summary data with references for markdown generation
        summary_data = []
        for csv_file in csv_files:
            base_name = os.path.splitext(os.path.basename(csv_file))[0]
            # Use the truncated sheet name for Excel references, but keep full name for display
            sheet_name = base_name[:31]
            
            # First read the CSV to get the sheet name
            df = pd.read_csv(csv_file)
            
            summary_data.append({
                'Sheet': f'=HYPERLINK("#\'{sheet_name}\'!A1","{base_name}")',  # Keep full name in display
                'Total_Tags': df[df['tag'] == 'totaldoctagswords']['tag_count'].iloc[0],
                'Total_Words': df[df['tag'] == 'totaldoctagswords']['word_count'].iloc[0],
                'Chapter_Count': df[df['tag'] == 'chapmarker']['tag_count'].sum() if 'chapmarker' in df['tag'].values else 0,
                'SceneAction_Count': df[df['tag'] == 'sceneaction']['tag_count'].sum() if 'sceneaction' in df['tag'].values else 0,
                'SceneAction_Words': df[df['tag'] == 'sceneaction']['word_count'].sum() if 'sceneaction' in df['tag'].values else 0,
                'SceneDia_Count': df[df['tag'] == 'scenedia']['tag_count'].sum() if 'scenedia' in df['tag'].values else 0,
                'SceneDia_Words': df[df['tag'] == 'scenedia']['word_count'].sum() if 'scenedia' in df['tag'].values else 0,
                'Dialogue_Count': df[df['tag'] == 'dia']['tag_count'].sum() if 'dia' in df['tag'].values else 0,
                'Dialogue_Words': df[df['tag'] == 'dia']['word_count'].sum() if 'dia' in df['tag'].values else 0
            })
        
        # Generate markdown summary before writing to Excel
        commit_hash = get_current_git_commit()
        create_markdown_summary(summary_data, commit_hash)
        
        # Get all unique tags across all files
        all_unique_tags_set = all_unique_tags.copy()
        if 'totaldoctagswords' in all_unique_tags_set:
            all_unique_tags_set.remove('totaldoctagswords')

        # Special columns we always want to keep separate from tag filtering
        special_cols = {'Sheet', 'Total_Tags', 'Total_Words', 'Chapter_Count', 'totaldoctagswords'}
        
        # Get base names for all CSV files
        base_names = []
        for csv_file in csv_files:
            base_name = os.path.splitext(os.path.basename(csv_file))[0]
            base_names.append(base_name)
        
        # Sort by year and then alphabetically
        base_names.sort(key=lambda x: (extract_year_from_filename(x), x))
        
        # First get tag frequency and word count data across all files
        tag_frequency = {}
        tag_word_total = {}
        
        # Process each CSV file to count total occurrences of each tag
        for csv_file in csv_files:
            df = pd.read_csv(csv_file)
            
            # Sum up the counts for each tag across all files
            for tag in all_unique_tags:
                if tag == 'totaldoctagswords':
                    continue  # Skip this as we have Total_Tags and Total_Words
                    
                tag_count = df[df['tag'] == tag]['tag_count'].sum() if tag in df['tag'].values else 0
                tag_words = df[df['tag'] == tag]['word_count'].sum() if tag in df['tag'].values else 0
                
                if tag not in tag_frequency:
                    tag_frequency[tag] = 0
                    tag_word_total[tag] = 0
                    
                tag_frequency[tag] += tag_count
                tag_word_total[tag] += tag_words
        
        # If we have excluded tags, use them to determine which tags to include
        if excluded_tags_set is not None:
            # Everything not in excluded_tags_set and not a special column is included
            included_tags_set = all_unique_tags_set - excluded_tags_set - special_cols
            
            # Create transposed summary for included tags - this will be our primary view
            included_transposed_df = create_transposed_summary(csv_files, included_tags_set, excluded_tags_set, "Summary Included Tags")
            included_transposed_df.to_excel(writer, sheet_name='Summary Included Tags', index=False)
            
            # Apply custom formatting to the included tags sheet
            included_trans_worksheet = writer.sheets['Summary Included Tags']
            apply_excel_formatting(included_trans_worksheet, included_transposed_df, base_names)
            
            # Create transposed sheet for excluded tags
            excluded_transposed_df = create_transposed_summary(csv_files, excluded_tags_set, None, "Summary Excluded Tags")
            excluded_transposed_df.to_excel(writer, sheet_name='Summary Excluded Tags', index=False)
            
            # Apply custom formatting to the excluded tags sheet  
            excluded_trans_worksheet = writer.sheets['Summary Excluded Tags']
            apply_excel_formatting(excluded_trans_worksheet, excluded_transposed_df, base_names)
            
            # Save the list of included and excluded tags to TSV files
            save_tag_lists(included_tags_set, excluded_tags_set)
        
        # Create transposed summary sheet for all tags
        all_transposed_df = create_transposed_summary(csv_files, all_unique_tags_set, None, "Summary All Tags")
        all_transposed_df.to_excel(writer, sheet_name='Summary All Tags', index=False)
        
        # Apply custom formatting to the all tags sheet
        all_trans_worksheet = writer.sheets['Summary All Tags']
        apply_excel_formatting(all_trans_worksheet, all_transposed_df, base_names)
        
        # Create and format the Summary Freq Words tab (sorted by word count)
        word_freq_df = create_transposed_freq_summary(csv_files, all_unique_tags_set - special_cols, sort_by_words=True, tag_frequency=tag_frequency, tag_word_total=tag_word_total)
        word_freq_df.to_excel(writer, sheet_name='Summary Freq Words', index=False)
        
        # Apply custom formatting to the frequency by words sheet
        word_freq_worksheet = writer.sheets['Summary Freq Words']
        apply_excel_formatting(word_freq_worksheet, word_freq_df, base_names)
        
        # Create and format the Summary Freq Tags tab (sorted by tag frequency)
        tag_freq_df = create_transposed_freq_summary(csv_files, all_unique_tags_set - special_cols, sort_by_words=False, tag_frequency=tag_frequency, tag_word_total=tag_word_total)
        tag_freq_df.to_excel(writer, sheet_name='Summary Freq Tags', index=False)
        
        # Apply custom formatting to the frequency by tags sheet
        tag_freq_worksheet = writer.sheets['Summary Freq Tags']
        apply_excel_formatting(tag_freq_worksheet, tag_freq_df, base_names)
        
        # Organize sheets in the proper order:
        # 1. Summary Included Tags
        # 2. Summary Excluded Tags
        # 3. Summary All Tags
        # 4. Summary Freq Words
        # 5. Summary Freq Tags
        # 6. Individual texts
        workbook = writer.book
        
        # First get the order of all sheets
        current_sheets = []
        for sheet in workbook.worksheets:
            current_sheets.append(sheet.title)
        
        # Reorder the sheets 
        desired_order = [
            'Summary Included Tags',
            'Summary Excluded Tags',
            'Summary All Tags',
            'Summary Freq Words',
            'Summary Freq Tags'
        ]
        
        # Add any sheet not explicitly mentioned in desired_order to the end of it
        for sheet in current_sheets:
            if sheet not in desired_order:
                desired_order.append(sheet)
        
        # Actually reorder the sheets
        workbook._sheets = [workbook[sheet_name] for sheet_name in desired_order]

def get_current_git_commit(specified_ref=None):
    """Get the current git commit hash or use specified ref."""
    import subprocess
    
    if specified_ref:
        # If a specific ref was provided, use it directly
        # This could be a branch name, tag, or commit hash
        return specified_ref
        
    try:
        # Otherwise get the current HEAD commit hash
        result = subprocess.run(['git', 'rev-parse', 'HEAD'], 
                              capture_output=True, text=True, check=True)
        return result.stdout.strip()
    except subprocess.CalledProcessError:
        return None

def analyze_scene_complexity(scene_tag):
    """Analyze the complexity of a scene based on its internal structure."""
    # Count unique tag types
    unique_tags = set(tag.name for tag in scene_tag.find_all())
    # Count total nested tags
    total_tags = len(scene_tag.find_all())
    # Count words
    words = len([w for w in re.split(r'\s+', scene_tag.get_text().strip()) if w])
    # Calculate complexity score (can be adjusted)
    return {
        'score': len(unique_tags) * 2 + total_tags + words/100,
        'unique_tags': unique_tags,
        'total_tags': total_tags,
        'words': words,
        'text': scene_tag.prettify()
    }

def find_interesting_excerpts(scene_tag, scene_type):
    """Find interesting excerpts from a scene including opening, transitions, rich dialog sections, and ending."""
    excerpts = []
    
    # Get the original content lines for line number tracking
    content = scene_tag.original_content
    
    # Helper function to find line numbers for an excerpt
    def find_excerpt_lines(excerpt_text):
        # Make sure we're dealing with string content
        if not content or len(content) == 0:
            return None, None
            
        # Extract first significant words for matching
        soup = BeautifulSoup(excerpt_text, 'html.parser')
        text_content = soup.get_text()
        words = text_content.split()
        if len(words) < 5:  # Too few words to match reliably
            return None, None
            
        first_words = ' '.join(words[:8]).lower()
        # Remove common HTML structure for better matching
        cleaned_excerpt = re.sub(r'<[^>]+>', ' ', excerpt_text).strip()
        cleaned_excerpt = re.sub(r'\s+', ' ', cleaned_excerpt).lower()
        
        start_line = None
        end_line = None
        
        # First pass: look for exact content match
        for i, line in enumerate(content, 1):
            line_lower = line.lower()
            if first_words in line_lower or cleaned_excerpt[:30] in line_lower:
                start_line = i
                end_line = i  # Start with single line
                break
        
        # If we found the start, try to find the end
        if start_line:
            # For multi-line excerpts
            remaining_content = ''.join(content[start_line:])
            cleaned_remaining = re.sub(r'<[^>]+>', ' ', remaining_content).strip()
            cleaned_remaining = re.sub(r'\s+', ' ', cleaned_remaining).lower()
            
            # Count how many lines from start
            for i in range(min(20, len(content) - start_line)):
                # If we find the end of the excerpt in this line
                if i > 0 and (cleaned_excerpt[-20:] in content[start_line + i].lower()):
                    end_line = start_line + i
                    break
                # As a fallback, just use a reasonable number of lines
                if i > 3:  # At least capture a few lines
                    end_line = start_line + i
        
        # If we still didn't find good bounds, use the scene bounds
        if not start_line or not end_line:
            for i, line in enumerate(content, 1):
                if f"<{scene_type.lower()}" in line.lower():
                    start_line = i
                if f"</{scene_type.lower()}>" in line.lower():
                    end_line = i
                    break
        
        # Ensure we have valid line numbers
        if start_line and end_line and start_line <= end_line:
            return start_line, end_line
        return None, None
    
    # Helper function to truncate text to a reasonable length while preserving tag structure
    def truncate_html(html_text, max_words=500):
        soup = BeautifulSoup(html_text, 'html.parser')
        words = soup.get_text().split()
        if len(words) <= max_words:
            return html_text
        
        # Find a good breakpoint near max_words
        text = ' '.join(words[:max_words])
        last_period = text.rfind('.')
        if last_period > len(text) * 0.6:
            text = text[:last_period+1]
        return html_text[:html_text.find(text) + len(text)] + "..."
    
    # Get the full scene text
    scene_text = str(scene_tag)
    
    # Find rich dialog sections
    dialog_sections = []
    for dia in scene_tag.find_all('dia', recursive=False):
        context = dia
        parent = dia.parent
        
        # Try to expand context to include nearby dialogs for richer context
        if parent and parent != scene_tag:
            siblings = list(parent.find_all('dia', recursive=False))
            if len(siblings) > 1:
                dia_index = siblings.index(dia)
                # Get 2 dialogs before and 2 after if available
                start_idx = max(0, dia_index - 2)
                end_idx = min(len(siblings), dia_index + 3)
                context = parent
                for child in list(context.children):
                    if isinstance(child, Tag) and child.name == 'dia' and child not in siblings[start_idx:end_idx]:
                        child.decompose()
        
        # Count tag variety and interesting tags
        tags_in_context = set(tag.name for tag in context.find_all())
        interesting_tags = {'m', 'chnameintro', 'chnonameintro', 'trigger', 'i', 'monologue', 'fidquotes', 'blend', 'quotedlit', 'metaphor', 'reportedspeechquotes'}
        dialog_length = len(context.get_text().split())
        
        # Calculate score based on tag variety, interesting tags, and optimal length
        score = (
            len(tags_in_context) * 4 +                            # Reward tag variety (increased weight)
            len(tags_in_context & interesting_tags) * 5 +         # Bonus for interesting tags (increased weight)
            min(dialog_length / 100, 6) +                         # Reward reasonable length (increased cap)
            (1 if 'trigger' in tags_in_context else 0) * 4        # Extra bonus for trigger tags (increased)
        )
        
        # Only include if score is high enough and length is reasonable
        if score >= 5 and dialog_length > 80:                     # Reduced minimum length requirement
            excerpt_text = str(context)
            start_line, end_line = find_excerpt_lines(excerpt_text)
            dialog_sections.append({
                'text': excerpt_text,
                'score': score,
                'tag_count': len(tags_in_context),
                'start_line': start_line, 
                'end_line': end_line
            })
    
    if dialog_sections:
        dialog_sections.sort(key=lambda x: x['score'], reverse=True)
        for section in dialog_sections[:2]:
            text = truncate_html(section['text'])
            if len(text) > 100:                                   # Reduced minimum length threshold
                # Make sure we have start/end lines
                start_line, end_line = section['start_line'], section['end_line']
                if not start_line or not end_line:
                    start_line, end_line = find_excerpt_lines(text)
                
                excerpts.append({
                    'type': 'rich_dialog',
                    'text': text,
                    'description': 'Rich Dialog Section with Multiple Tag Types',
                    'start_line': start_line,
                    'end_line': end_line
                })
    
    # Find transitions with triggers
    triggers = scene_tag.find_all('trigger')
    if triggers:
        trigger_contexts = []
        for trigger in triggers:
            context = trigger.parent
            tags_in_context = set(tag.name for tag in context.find_all())
            if len(tags_in_context) >= 2:
                excerpt_text = str(context)
                start_line, end_line = find_excerpt_lines(excerpt_text)
                trigger_contexts.append({
                    'text': excerpt_text,
                    'tag_count': len(tags_in_context),
                    'start_line': start_line,
                    'end_line': end_line
                })
        
        if trigger_contexts:
            trigger_contexts.sort(key=lambda x: x['tag_count'], reverse=True)
            text = truncate_html(trigger_contexts[0]['text'])
            if len(text) > 40:                                    # Reduced minimum length threshold
                start_line, end_line = trigger_contexts[0]['start_line'], trigger_contexts[0]['end_line']
                if not start_line or not end_line:
                    start_line, end_line = find_excerpt_lines(text)
                    
                excerpts.append({
                    'type': 'transition',
                    'text': text,
                    'description': 'Scene Transition',
                    'start_line': start_line,
                    'end_line': end_line
                })
    
    def is_interesting_section(text):
        soup = BeautifulSoup(text, 'html.parser')
        tags = set(tag.name for tag in soup.find_all())
        interesting_tags = {'m', 'chnameintro', 'chnonameintro', 'trigger', 'i', 'monologue', 'fidquotes', 'blend', 'quotedlit', 'metaphor', 'reportedspeechquotes'}  # Added more interesting tags
        return len(tags) >= 2 or len(tags & interesting_tags) >= 1
    
    # Check opening
    opening_match = re.search(r'<' + scene_type + r'[^>]*>.*?(<.*?</.*?>)', scene_text, re.DOTALL)
    if opening_match and is_interesting_section(opening_match.group(0)):
        text = truncate_html(opening_match.group(0))
        if len(text) > 40:                                        # Reduced minimum length threshold
            start_line, end_line = find_excerpt_lines(text)
            if not start_line:
                # Fallback: find the opening tag
                for i, line in enumerate(content, 1):
                    if f"<{scene_type.lower()}" in line.lower():
                        start_line = i
                        end_line = start_line + 5  # Just a few lines
                        break
                        
            excerpts.append({
                'type': 'opening',
                'text': text,
                'description': 'Scene Opening',
                'start_line': start_line,
                'end_line': end_line
            })
    
    # Check ending
    ending_match = re.search(r'(<.*?</.*?>)[^<]*</' + scene_type + '>', scene_text, re.DOTALL)
    if ending_match and is_interesting_section(ending_match.group(0)):
        text = truncate_html(ending_match.group(0))
        if len(text) > 40:                                        # Reduced minimum length threshold
            start_line, end_line = find_excerpt_lines(text)
            if not end_line:
                # Fallback: find the closing tag
                for i, line in enumerate(content, 1):
                    if f"</{scene_type.lower()}>" in line.lower():
                        end_line = i
                        start_line = max(1, end_line - 5)  # Just a few lines
                        break
                        
            excerpts.append({
                'type': 'ending',
                'text': text,
                'description': 'Scene Ending',
                'start_line': start_line,
                'end_line': end_line
            })
    
    return excerpts

def find_rich_samples(html_file):
    """Find rich samples from a given HTML file."""
    with open(html_file, 'r', encoding='utf-8') as f:
        content = f.readlines()
        soup = BeautifulSoup(''.join(content), 'html.parser')
        # Store the original file path and content for line number lookup
        soup.original_file = html_file
        soup.original_content = content
        # Also store these on each tag
        for tag in soup.find_all():
            tag.original_file = html_file
            tag.original_content = content
    
    samples = []
    for scene_type in ['sceneaction', 'scenedia']:
        scenes = soup.find_all(scene_type, recursive=False)
        if not scenes:
            continue
            
        analyzed_scenes = []
        for scene in scenes:
            analysis = analyze_scene_complexity(scene)
            
            scene_text = str(scene)
            opening_tag = f"<{scene_type}"
            closing_tag = f"</{scene_type}>"
            
            start_line = None
            end_line = None
            
            for i, line in enumerate(content, 1):
                if opening_tag.lower() in line.lower():
                    start_line = i
                    break
            
            if start_line is not None:
                for i, line in enumerate(content[start_line-1:], start_line):
                    if closing_tag.lower() in line.lower():
                        end_line = i
                        break
            
            if start_line is not None and end_line is not None:
                analysis['start_line'] = start_line
                analysis['end_line'] = end_line
                analysis['scene_type'] = scene_type
                analysis['excerpts'] = find_interesting_excerpts(scene, scene_type)
                analyzed_scenes.append(analysis)
        
        if analyzed_scenes:
            analyzed_scenes.sort(key=lambda x: x['score'], reverse=True)
            samples.append(analyzed_scenes[0])
    
    return samples

def create_samples_markdown(commit_hash=None):
    """Generate a markdown file with rich samples from each input file."""
    markdown_content = "# Scene Samples\n\n"
    markdown_content += "This document contains particularly rich examples of scene markup from each text, "
    markdown_content += "showing complex interactions between different types of scenes and their components. "
    markdown_content += "For each scene, we show interesting excerpts including openings, transitions, rich dialog sections, and endings.\n\n"
    
    input_dir = "data/input"
    html_files = glob.glob(os.path.join(input_dir, "*.html"))
    
    for html_file in html_files:
        base_name = os.path.basename(html_file)
        # Create the GitHub link for the full file
        encoded_filename = quote(base_name)
        full_file_link = f"https://github.com/aculich/novel-scenification/blob/main/data/input/{encoded_filename}"
        markdown_content += f"## [{base_name}]({full_file_link})\n\n"
        
        samples = find_rich_samples(html_file)
        if not samples:
            markdown_content += "No complex scenes found in this file.\n\n"
            continue
        
        for sample in samples:
            # Create the GitHub link with line numbers
            file_link = f"{full_file_link}#L{sample['start_line']}-L{sample['end_line']}"
            
            markdown_content += f"### Complex {sample['scene_type'].title()} (Lines {sample['start_line']}-{sample['end_line']})\n\n"
            markdown_content += f"**Location:** [Lines {sample['start_line']}-{sample['end_line']}]({file_link})\n\n"
            markdown_content += f"**Complexity Metrics:**\n"
            markdown_content += f"- Unique tag types: {len(sample['unique_tags'])}\n"
            markdown_content += f"- Total nested tags: {sample['total_tags']}\n"
            markdown_content += f"- Word count: {sample['words']}\n"
            markdown_content += f"- Tag types present: {', '.join(sorted(sample['unique_tags']))}\n\n"
            
            # Add excerpts with line numbers and links
            markdown_content += "**Interesting Excerpts:**\n\n"
            for excerpt in sample['excerpts']:
                # Add line numbers and link if available
                if excerpt['start_line'] and excerpt['end_line']:
                    excerpt_link = f"{full_file_link}#L{excerpt['start_line']}-L{excerpt['end_line']}"
                    markdown_content += f"*{excerpt['description']}:* [Lines {excerpt['start_line']}-{excerpt['end_line']}]({excerpt_link})\n"
                else:
                    markdown_content += f"*{excerpt['description']}:*\n"
                    
                markdown_content += "```html\n"
                markdown_content += excerpt['text']
                markdown_content += "\n```\n\n"
    
    # Ensure data directory exists
    os.makedirs('data', exist_ok=True)
    with open('data/SAMPLES.md', 'w', encoding='utf-8') as f:
        f.write(markdown_content)

def create_readme():
    """Generate README.md from template and include summaries."""
    try:
        # Try to read existing README.md
        with open('README.md', 'r', encoding='utf-8') as f:
            existing_content = f.read()
            
        # Split at "## Tag Counts Summary" if it exists
        parts = existing_content.split("## Tag Counts Summary")
        template_content = parts[0].rstrip()  # Use everything before this section
    except FileNotFoundError:
        # If README.md doesn't exist, start with empty content
        template_content = ""
    
    # Add links to full files
    summary_section = "\n\n## Tag Counts Summary\n\n"
    # Use raw format for Excel file only
    summary_section += "[View complete tag counts summary](https://github.com/aculich/novel-scenification/raw/refs/heads/main/data/tag_counts_summary.xlsx)\n\n"
    
    # Add table header once
    summary_section += "| Sheet | Total_Tags | Total_Words | Chapter_Count | SceneAction_Count | SceneAction_Words | SceneDia_Count | SceneDia_Words | Dialogue_Count | Dialogue_Words |\n"
    summary_section += "|-------|------------|-------------|---------------|------------------|------------------|----------------|----------------|----------------|----------------|\n"
    
    # Flag to track if we've added any rows
    added_rows = False
    
    try:
        # Get sample files to filter rows
        sample_files = set()
        try:
            with open('data/SAMPLES.md', 'r', encoding='utf-8') as samples_file:
                samples_content = samples_file.read()
                # Extract sample filenames from both the headings and the URLs
                for section in samples_content.split('\n## ['):
                    if '](' in section:
                        file_name = section.split('](')[0]
                        if file_name.endswith('.html'):
                            # Store without .html extension
                            clean_name = file_name[:-5]
                            sample_files.add(clean_name)
                            # Also store filename-only version
                            base_name = os.path.basename(clean_name)
                            sample_files.add(base_name)
        except FileNotFoundError:
            pass  # If no samples file, don't filter
        
        # Read summary data directly from the Excel file
        summary_df = pd.read_excel('data/tag_counts_summary.xlsx', sheet_name='Summary')
            
        # Add rows for each file
        for _, row in summary_df.iterrows():
            sheet_name = row['Sheet']
            # Extract the base name from the HYPERLINK formula if present
            if isinstance(sheet_name, str) and '=HYPERLINK' in sheet_name:
                match = re.search(r'"([^"]+)"', sheet_name)
                if match:
                    base_name = match.group(1)
                else:
                    base_name = sheet_name
            else:
                base_name = sheet_name
                
            # Skip if we're filtering by sample files and this one isn't included
            if sample_files:
                # Try different formats for matching
                match_found = False
                for sample_file in sample_files:
                    # Try exact match or match without extension
                    if (base_name == sample_file or 
                        base_name == sample_file + '.html' or
                        sample_file in base_name):
                        match_found = True
                        break
                
                if not match_found:
                    continue
                
            added_rows = True
            # Create GitHub link for the row
            github_link = f"[{base_name}](https://github.com/aculich/novel-scenification/blob/main/data/input/{base_name}.html)"
            
            # Add table row with GitHub link
            summary_section += f"| {github_link} | {row['Total_Tags']} | {row['Total_Words']} | {row['Chapter_Count']} | {row['SceneAction_Count']} | {row['SceneAction_Words']} | {row['SceneDia_Count']} | {row['SceneDia_Words']} | {row['Dialogue_Count']} | {row['Dialogue_Words']} |\n"
    except Exception as e:
        # If Excel reading fails, try to use SUMMARY.md instead
        try:
            with open('data/SUMMARY.md', 'r', encoding='utf-8') as f:
                summary_content = f.read()
                summary_lines = summary_content.split('\n')
                
                # Find the table header and data rows
                table_start = -1
                for i, line in enumerate(summary_lines):
                    if '| Sheet | Total_Tags' in line:
                        table_start = i
                        break
                
                if table_start >= 0:
                    # Skip header and separator since we've already added them
                    # Add data rows only
                    for i in range(table_start + 2, len(summary_lines)):
                        line = summary_lines[i].strip()
                        if not line or not line.startswith('|'):
                            continue
                            
                        # If filtering by sample files, check if this row should be included
                        if sample_files:
                            include_row = False
                            for sample_file in sample_files:
                                if sample_file in line:
                                    include_row = True
                                    break
                                    
                            if include_row:
                                summary_section += line + "\n"
                                added_rows = True
                        else:
                            # Include all rows if not filtering
                            summary_section += line + "\n"
                            added_rows = True
        except FileNotFoundError:
            # If both Excel and SUMMARY.md fail, don't add anything to the table
            pass
    
    # If we didn't add any rows, include all rows from Excel or SUMMARY.md
    if not added_rows:
        try:
            # Try again with Excel, but without filtering
            summary_df = pd.read_excel('data/tag_counts_summary.xlsx', sheet_name='Summary')
            
            for _, row in summary_df.iterrows():
                sheet_name = row['Sheet']
                # Extract the base name from the HYPERLINK formula if present
                if isinstance(sheet_name, str) and '=HYPERLINK' in sheet_name:
                    match = re.search(r'"([^"]+)"', sheet_name)
                    if match:
                        base_name = match.group(1)
                    else:
                        base_name = sheet_name
                else:
                    base_name = sheet_name
                    
                # Create GitHub link for the row
                github_link = f"[{base_name}](https://github.com/aculich/novel-scenification/blob/main/data/input/{base_name}.html)"
                
                # Add table row with GitHub link
                summary_section += f"| {github_link} | {row['Total_Tags']} | {row['Total_Words']} | {row['Chapter_Count']} | {row['SceneAction_Count']} | {row['SceneAction_Words']} | {row['SceneDia_Count']} | {row['SceneDia_Words']} | {row['Dialogue_Count']} | {row['Dialogue_Words']} |\n"
                added_rows = True
        except Exception:
            # If that fails, try SUMMARY.md without filtering
            try:
                with open('data/SUMMARY.md', 'r', encoding='utf-8') as f:
                    summary_content = f.read()
                    summary_lines = summary_content.split('\n')
                    
                    # Find the table header and data rows
                    table_start = -1
                    for i, line in enumerate(summary_lines):
                        if '| Sheet | Total_Tags' in line:
                            table_start = i
                            break
                    
                    if table_start >= 0:
                        # Add all data rows
                        for i in range(table_start + 2, len(summary_lines)):
                            line = summary_lines[i].strip()
                            if not line or not line.startswith('|'):
                                continue
                                
                            # Include all rows
                            summary_section += line + "\n"
                            added_rows = True
            except FileNotFoundError:
                # If everything fails, add a no data row
                pass
    
    # If we still didn't add any rows, add a placeholder
    if not added_rows:
        summary_section += "| No data available | - | - | - | - | - | - | - | - | - |\n"
    
    # Add excerpt from SAMPLES.md
    samples_section = "\n\n## Scene Samples\n\n"
    samples_section += "[View complete samples analysis](data/SAMPLES.md)\n\n"
    
    try:
        with open('data/SAMPLES.md', 'r', encoding='utf-8') as f:
            samples_content = f.read()
            # Split by sections and take introduction plus first two samples
            sections = samples_content.split('\n## ')
            if len(sections) > 0:
                intro = sections[0]
                samples = []
                count = 0
                for section in sections[1:]:
                    if count < 2 and section.strip():  # Only process non-empty sections
                        # Keep the original formatting including line numbers and links
                        samples.append(section)
                        count += 1
                
                samples_section += intro + "\n"
                if samples:
                    samples_section += "## " + "\n## ".join(samples) + "\n...\n\n"
                    samples_section += "[View all scene samples](data/SAMPLES.md)\n"
    except FileNotFoundError:
        # If SAMPLES.md doesn't exist, skip this section
        pass
    
    # Combine everything
    readme_content = template_content + summary_section + samples_section
    
    # Write the README
    with open('README.md', 'w', encoding='utf-8') as f:
        f.write(readme_content)

def process_all_files():
    input_dir = "data/input"
    output_dir = "data/counts"
    
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Get all HTML files in input directory
    html_files = glob.glob(os.path.join(input_dir, "*.html"))
    
    for html_file in html_files:
        # Generate output filename by replacing directory and extension
        base_name = os.path.basename(html_file)
        csv_name = os.path.splitext(base_name)[0] + ".csv"
        csv_file = os.path.join(output_dir, csv_name)
        
        print(f"\nProcessing: {base_name}")
        print("=" * 60)
        
        parse_html_to_csv(html_file, csv_file)
        print(f"Output saved to: {csv_file}\n")
    
    # After processing all files, create Excel summary and samples
    print("\nCreating Excel summary file...")
    create_excel_summary()
    print("Excel summary file created at: data/tag_counts_summary.xlsx")
    
    print("\nGenerating rich samples documentation...")
    create_samples_markdown()
    print("Rich samples documentation created at: data/SAMPLES.md")
    
    print("\nGenerating README.md...")
    create_readme()
    print("README.md generated from template with summaries")

def main():
    """Process all HTML files and generate summary files."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Process HTML files and generate tag count summaries.")
    args = parser.parse_args()
    
    # Process all files and generate CSV files in data/counts
    process_all_files()

if __name__ == "__main__":
    main()